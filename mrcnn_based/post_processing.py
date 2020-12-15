#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth, Tobias Kapetanopoulos"
__copyright__ = "Copyright 2020, M. Hahn-Klimroth, T. Kapetanopoulos, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "MIT"
__version__ = "1.0"
__status__ = "Development"

import configuration as cf
import copy
import csv, os
import numpy as np
from openpyxl import Workbook
import matplotlib.pyplot as plt
from matplotlib.collections import PolyCollection


def ensure_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def post_process_night(single_frame_csv, joint_interval_csv, 
                       individual_code, datum, position_files,
                       output_folder_prediction = cf.FINAL_STORAGE_PREDICTION_FILES, 
                       behavior_mapping = cf.BEHAVIOR_MAPPING,
                       is_test = False, extension = ''):
    
    
    def _map_behaviors( phase_list, behavior_mapping = behavior_mapping ):
        """
        Maps phases, e.g. if we want to join lying and sleeping.
        """
        for phase in phase_list:
            phase[1] = behavior_mapping[phase[1]]        
        return phase_list
    
    def _apply_rolling_average( list_of_dist, amount_back, weights):
        rolling_avg_per_img = []

        for j in range(len(list_of_dist)):
            go_back_here = min(j, amount_back)        
            curr_dist = np.array([0.0]*len(cf.BEHAVIOR_NAMES))
            
            for k in range(j - go_back_here, j+1):
                curr_dist += list_of_dist[k]*weights
            curr_dist = curr_dist / np.linalg.norm(curr_dist, ord=1)
            rolling_avg_per_img.append(curr_dist)
        
        return rolling_avg_per_img
    
    def _read_csv(filepath):
        """ 
            Requires path to csv name || start || end || standing || lying || sleeping || out 
            Outputs [ [standing, lying, sleeping, out, truncated] ]
        """
        
        ret = []
        
        with open(filepath, "r") as f:

            csv_real = csv.reader(f, delimiter=',')
            j = 0    
            for row in csv_real:
                j += 1
                if j == 1:
                    continue
                ret.append( [ np.float(row[3]), np.float(row[4]), np.float(row[5]), np.float(row[6]), 0.0 ] )
        
        return ret   
    
    def _single_frame_prediction_to_intervalprediction(dist_list, n = cf.INTERVAL_LENGTH):
        list_of_intervals = [dist_list[i * n:(i + 1) * n] for i in range((len(dist_list) + n - 1) // n )]
        ret = []
        for interval in list_of_intervals:
            curr = np.array( [0.0]*len(cf.BEHAVIOR_NAMES) )
            for img_dist in interval:
                curr += np.array( img_dist )
            curr = curr / np.linalg.norm(curr, ord=1)
            
            ret.append(curr)
        return ret
    
    def _calculate_joint_prediction( dist1, dist2, weights = cf.WEIGHTS_POSTPROCESS, num_behav = 5, is_test = is_test ):
        """
        
    
        Parameters
        ----------
        dist 1: joint images
        dist 2: single frames
        
    
        Returns
        -------
        
            Output: List of intervals, for each 
    
        """
        
        if is_test:
            print("Testing mode on real data.")
            return dist1
        
        len1 = len(dist1)
        len2 = len(dist2)
        
        ret = []
        
        if len1 != len2:
            print("Warning: Different lenghts of predictions.", len1, len2)
        
        minlen = min(len1, len2)
        
        for j in range(minlen):
            pred1 = dist1[j]
            pred2 = dist2[j]
            
            pred = weights[0] * np.array(pred1) + weights[1]*np.array(pred2)
            
            ret.append(pred)
        
        return ret
    
    def _write_prediction_csv(interval_dist, output_folder, filename):
        file_path = output_folder + filename
        ensure_dir(output_folder)
        
        with open(file_path, mode='w+') as csv_out:
            csv_write = csv.writer(csv_out, delimiter=",")
            csv_write.writerow( ["Interval", "Startframe", "Endframe", 
                                 cf.BEHAVIOR_NAMES[0],
                                 cf.BEHAVIOR_NAMES[1],
                                 cf.BEHAVIOR_NAMES[2],
                                 cf.BEHAVIOR_NAMES[3],
                                 cf.BEHAVIOR_NAMES[4]] )
            
            h = 1
            for behav_int in interval_dist:
                sf = (h-1)*cf.INTERVAL_LENGTH + 1
                ef = h*cf.INTERVAL_LENGTH
                row = [h, sf, ef, behav_int[0], behav_int[1], behav_int[2] ,behav_int[3], behav_int[4]]
                csv_write.writerow(row)
                h += 1
                
    def _sparse_encoding(interval_dist):
        ret = []
        for interval in interval_dist:
            ret.append( np.argmax( interval ) )
        
        return ret
    
    def _extract_single_phases(np_array, timeinterval = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        np_array : TYPE
            DESCRIPTION.
        timeinterval : TYPE, optional
            DESCRIPTION. The default is cf.INTERVAL_LENGTH.
    
        Returns
        -------
        phases : [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]
    
        """
        phases = []
        LastBehav = np_array[0]
        iCurrLen = 1
        j = 1
        start_interval = 1

        for DistLine in np_array:
            
            if LastBehav == DistLine:
                iCurrLen += 1
            else:
                phases.append([iCurrLen*timeinterval, LastBehav, start_interval, j-1])
                iCurrLen = 1
                start_interval = j
                LastBehav = DistLine
            j += 1

        phases.append([iCurrLen*timeinterval, LastBehav, phases[-1][3] + 1, phases[-1][3] + 1 + iCurrLen])
        return phases
    
    
    def _join_phases(phase_list):
        """        
            Joins consecutive phases of the same behavior to a single phase.

        """
    
        ret = [phase_list[0]]
        
        for j in range(1, len(phase_list)):
            if phase_list[j][1] == phase_list[j-1][1]:
                ret[-1][0] += phase_list[j][0]
                ret[-1][3] = phase_list[j][3]
            else:
                ret.append(phase_list[j])
        
        return ret
    
    
    def _remove_short_phases(phase_list, interval_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Removes those short phases which are very unlikely due to configuration
        """
        
  
        phase_list = _join_phases(phase_list)
        
        if len(phase_list) < 3:
            return phase_list
        
        
        changes_done = True
        
        while changes_done:
        
            last_behavior = phase_list[0][1]
            current_behavior = None
            next_behavior = None
            
            changes_done = False
            
            
            for j in range(1, len(phase_list)-1):
                last_behavior = phase_list[j-1][1]
                current_behavior = phase_list[j][1]
                next_behavior = phase_list[j+1][1]
                
                # lying
                if [last_behavior, current_behavior, next_behavior] == [2,1,2]:
                    if phase_list[j][0] < cf.MIN_LEN_SLS*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,0]:
                    if phase_list[j][0] < cf.MIN_LEN_ALA*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,1,2]:
                    if phase_list[j][0] < cf.MIN_LEN_ALS*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,1,0]:
                    if phase_list[j][0] < cf.MIN_LEN_SLA*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                        
                        
                # standing
                elif [last_behavior, current_behavior, next_behavior] == [2,0,2]:
                    if phase_list[j][0] < cf.MIN_LEN_SAS*interval_len:
                        phase_list[j][1] = 2
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,1]: 
                    if phase_list[j][0] < cf.MIN_LEN_LAL*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,0,2]:
                    if phase_list[j][0] < cf.MIN_LEN_LAS*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [2,0,1]:
                    if phase_list[j][0] < cf.MIN_LEN_SAL*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                        
                # sleeping
                elif [last_behavior, current_behavior, next_behavior] == [0,2,0]: 
                    if phase_list[j][0] < cf.MIN_LEN_ASA*interval_len:
                        phase_list[j][1] = 0
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [0,2,1]: 
                    if phase_list[j][0] < cf.MIN_LEN_ASL*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,1]: 
                    if phase_list[j][0] < cf.MIN_LEN_LSL*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                elif [last_behavior, current_behavior, next_behavior] == [1,2,0]: 
                    if phase_list[j][0] < cf.MIN_LEN_LSA*interval_len:
                        phase_list[j][1] = 1
                        changes_done = True
                        
                # truncation
                elif current_behavior == 4: 
                    if phase_list[j][0] < cf.MIN_LEN_TRUNCATION*interval_len:
                        phase_list[j][1] = last_behavior
                        changes_done = True         
        
            phase_list = _join_phases(phase_list)
            
        return phase_list

    def _remove_out(phase_list):
        """
        Removes those phases of behavior out which are shorter than the value given by the configuration
        """
        j = 0
        for phase in phase_list:
            if phase[1] == 3 and j >= 1: # it is out
                if phase[0] <= cf.MIN_LEN_OUT*cf.INTERVAL_LENGTH and phase[2] > cf.MIN_TIME_OUT:
                    phase[1] = phase_list[j-1][1]
            j += 1
        
        return phase_list
    
    def _remove_truncated_images(phase_list):
        """
        Removes those phases of truncation out which are longer than the value given by the configuration
        """
        j = 0
        for phase in phase_list:
            if phase[1] == 4 and j >= 1: # it is truncated
                if phase[0] >= cf.MIN_LEN_TRUNCATION_SWAP*cf.INTERVAL_LENGTH:
                    phase[1] = cf.TRUNCATION_REAL_BEHAVIOR_LONG
            j += 1
        
        return phase_list
    
    
    def _get_time(interval_num, video_start = cf.VIDEO_START_TIME, interval_len = cf.INTERVAL_LENGTH):
        """
        
    
        Parameters
        ----------
        interval_num : TYPE
            DESCRIPTION.
        video_start : TYPE, optional
            DESCRIPTION. The default is VIDEO_START_TIME.
    
        Returns
        -------
        startzeit, endzeit, startframe, endframe
    
        """
        startframe = (interval_num-1)*interval_len + 1
        endframe = interval_num*interval_len
        
        start_hours = startframe // (60*60)
        start_minutes = (startframe % (60*60)) // 60
        start_seconds = startframe % 60
        
        end_hours = endframe // (60*60)
        end_minutes = (endframe % (60*60)) // 60
        end_seconds = endframe % 60
        
        start_time = str((video_start + start_hours)%24).zfill(2) + ":" + str(start_minutes).zfill(2) + ":" + str(start_seconds).zfill(2)
        end_time = str((video_start + end_hours)%24).zfill(2) + ":" + str(end_minutes).zfill(2) + ":" + str(end_seconds).zfill(2)
        
        return start_time, end_time, startframe, endframe
    
    def _get_sparse_intervals_from_phases(phases):
       
        new_intervallist = [] # sequence of behaviors
        for j in range(1, len(phases)+1):
            curr_phase = phases[j-1]
            curr_start = curr_phase[2]
            curr_end = curr_phase[3]
            curr_dur = curr_end - curr_start
            for i in range(curr_dur + 1):
                new_intervallist.append(curr_phase[1])
        return new_intervallist
     
    
    def _write_xlsx_statistics(phases_ordered, phases_unordered, interval_list, outputfolder, 
                               filename, names_behav = cf.BEHAVIOR_NAMES, extension=extension):
        wb = Workbook()
        ws = wb.active
        ws.title = "Zeitintervalle"
        
        ws["A1"] = "Zeitintervall"
        ws["B1"] = "Startzeit"
        ws["C1"] = "Endzeit"
        ws["D1"] = "Startframe"
        ws["E1"] = "Endframe"
        ws["F1"] = "Verhaltenscode"
        ws["G1"] = "Verhaltenswort"
        
        
        for j in range(len(interval_list)):
            starttime, endtime, startframe, endframe = _get_time(j+1)
            ws["A"+str(j+2)] = j+1
            ws["B"+str(j+2)] = starttime
            ws["C"+str(j+2)] = endtime
            ws["D"+str(j+2)] = startframe
            ws["E"+str(j+2)] = endframe
            ws["F"+str(j+2)] = interval_list[j]
            ws["G"+str(j+2)] = names_behav[interval_list[j]]
            
        
                    
        #  [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ]
        # phases: [ [dauer (sek), verhaltenscode, startinterval, endinterval] ] (stat sheet 2)
        ws2 = wb.create_sheet("Aktivitätsphasen_geordnet", -1)
        
        ws2["A1"] = "Phase"
        ws2["B1"] = "Nummer der Phase des Verhaltens"
        ws2["C1"] = "Startzeit"
        ws2["D1"] = "Endzeit"
        ws2["E1"] = "Startintervall"
        ws2["F1"] = "Endintervall"
        ws2["G1"] = "Länge [sec]"
        ws2["H1"] = "Verhaltenscode"
        ws2["I1"] = "Verhaltenswort"
        
        j = 0
        amount_phases = [0]*len(names_behav)
        for phase in phases_unordered:
            amount_phases[phase[1]] += 1
            starttime, _, _, _ = _get_time(phase[2])
            _, endtime, _, _ = _get_time(phase[3])
            
            ws2["A"+str(j+2)] = j+1
            ws2["B"+str(j+2)] = amount_phases[phase[1]] 
            ws2["C"+str(j+2)] = starttime
            ws2["D"+str(j+2)] = endtime
            ws2["E"+str(j+2)] = phase[2]
            ws2["F"+str(j+2)] = phase[3]
            ws2["G"+str(j+2)] = phase[0]
            ws2["H"+str(j+2)] = phase[1]
            ws2["I"+str(j+2)] = names_behav[phase[1]]
            
            j += 1
            
            
        # phases_ordered [ [ [dauer (s), 0, startinterval, endinterval] ... ], [dauer (s), 1, startinterval, endinterval], ... ] (stat sheet 3) 
        ws3 = wb.create_sheet("Aktivitätsphasen_Übersicht", -1)
        colname=["A", "B", "C", "D", "E", "F"]
        
        for i in range(4):
            ws3[colname[i]+"1"] = names_behav[i]
            j = 0
            for phase_stand in phases_ordered[i]:
                ws3[colname[i]+str(j+2)] = phase_stand
                j += 1
        
        
        
        
        ws4 = wb.create_sheet("Statistik", -1)
        
        ws4["A2"] = "Anzahl Phasen"
        ws4["A3"] = "Gesamtdauer [sec]"
        ws4["A4"] = "Anteil [%]"
        
        ws4["A6"] = "Median [sec]"
        ws4["A7"] = "0.25-Quantil"
        ws4["A8"] = "0.75-Quantil"
        ws4["A10"] = "Mean [sec]"
        ws4["A11"] = "SEM"
        
        total_duration = 0
        for j in range(len(names_behav)):
            total_duration += np.sum(phases_ordered[j])
        for j in range(len(names_behav)):
            amount = len(phases_ordered[j])
            ws4[colname[j+1]+"1"] = names_behav[j]
            if amount > 0:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = np.sum(phases_ordered[j])
                ws4[colname[j+1]+"4"] = round(100 * np.sum(phases_ordered[j]) / total_duration, 1)
                
                ws4[colname[j+1]+"6"] = np.median(phases_ordered[j])
                ws4[colname[j+1]+"7"] = np.quantile(phases_ordered[j], 0.25)
                ws4[colname[j+1]+"8"] = np.quantile(phases_ordered[j], 0.75)
                ws4[colname[j+1]+"10"] = np.mean(phases_ordered[j])
                ws4[colname[j+1]+"11"] = np.std(phases_ordered[j]) / np.sqrt(amount)
            else:
                ws4[colname[j+1]+"2"] = amount
                ws4[colname[j+1]+"3"] = 0
                ws4[colname[j+1]+"4"] = 0
                
                ws4[colname[j+1]+"6"] = 0
                ws4[colname[j+1]+"7"] = 0
                ws4[colname[j+1]+"8"] = 0
                ws4[colname[j+1]+"10"] = 0
                ws4[colname[j+1]+"11"] = 0
        
        ensure_dir(outputfolder)
        wb.save(outputfolder + filename + extension + '.xlsx')
    
    def _draw_timeline(phases_unordered, save_path, title, behav_names2 = cf.BEHAVIOR_NAMES, 
                       colormapping = cf.COLOR_MAPPING, extension = extension):
        
        behav_names = copy.deepcopy(behav_names2)
        data = []
        for phase in phases_unordered:
            startframe = phase[2]
            endframe = phase[3]
            behavior = behav_names[phase[1]]
            
            data.append((startframe, endframe, behavior))
        
        cats ={}
        for j in range(len(behav_names)):
            cats[behav_names[j]] = (j+1)/2
        
        verts = []
        colors = []
        for d in data:
            v =  [(d[0], cats[d[2]]-.075),
                  (d[0], cats[d[2]]+.075),
                  (d[1], cats[d[2]]+.075),
                  (d[1], cats[d[2]]-.075),
                  (d[0], cats[d[2]]-.075)]
            verts.append(v)
            colors.append(colormapping[d[2]])
        
        bars = PolyCollection(verts, facecolors=colors)
        
        x_ticks_set = []
        x_ticks_labels = []
        for j in range( int(cf.VIDEO_LENGTH / cf.INTERVAL_LENGTH) - 1):
            if j % int(3600/cf.INTERVAL_LENGTH) == 0:
                start_time, end_time, startframe, endframe = _get_time(j)
                hh, mm, ss = start_time.split(":")
                hh = str(int(hh)+1).zfill(2)
                mm = ':00'
                
                x_ticks_labels.append(hh + mm)
                x_ticks_set.append(j)
            
        
        
        
        fig = plt.figure(figsize=(13,4))
        ax = fig.add_subplot(111)
        
        fig.autofmt_xdate(rotation=60)
        fig.suptitle(title, fontsize=16)
        
        ax.add_collection(bars)
        ax.autoscale()
        #loc = mdates.MinuteLocator(byminute=[0,15,30,45])
        #ax.xaxis.set_major_locator(loc)
        #ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(loc))
        
        ax.set_xticks( x_ticks_set )
        ax.set_xticklabels( x_ticks_labels )
        ax.set_yticks([0.5, 1, 1.5, 2, 2.5, 3])
        behav_names.append(" ")
        ax.set_yticklabels(behav_names)
        
      
        plt.savefig(save_path + extension + '.png')
        plt.close()
    
    def _mark_truncated_images_single(single_frame_dist, xmin = cf.TRUNCATION_X_MIN, 
                                      ymin = cf.TRUNCATION_Y_MIN, 
                                      position_files = position_files):

        ret = []
        for j in range(len(single_frame_dist)):            
            if np.argmax(single_frame_dist[j]) == 3:
                ret.append(single_frame_dist[j])
                continue
            
            time_interval = str((j // cf.INTERVAL_LENGTH) + 1).zfill(7)
            pos_info = time_interval + '.txt'
            if not os.path.exists(position_files + pos_info):
                ret.append(single_frame_dist[j])
                continue

            info_file = open(position_files + pos_info, 'r')
            info_content = info_file.read().split('\n')
            
            found_img = False
            index = 0
            
            for box_info in info_content:
                if box_info.startswith( str(j).zfill(7) ):
                    found_img = True
                    break
                index += 1
            
            if found_img:
                coordinates = info_content[index].split('-')[1].split('*')
                y1, x1 = int(coordinates[0]), int(coordinates[1])
                
                if x1 > xmin or y1 > ymin:
                    ret.append( [0.0, 0.0, 0.0, 0.0, 1.0] )
                else:
                    ret.append( single_frame_dist[j]  )
            else:
                ret.append( single_frame_dist[j]  )

        return ret
    
    def _mark_truncated_images_joint(joint_frame_dist, xmin = cf.TRUNCATION_X_MIN, 
                                      ymin = cf.TRUNCATION_Y_MIN, 
                                      position_files = position_files):
        #print(len(joint_frame_dist))
        ret = []
        for j in range(len(joint_frame_dist)):            
            if np.argmax(joint_frame_dist[j]) == 3:
                ret.append(joint_frame_dist[j])
                continue
            
            time_interval = str( j + 1 ).zfill(7)
            pos_info = time_interval + '.txt'
            if not os.path.exists(position_files + pos_info):
                ret.append(single_frame_dist[j])
                continue

            info_file = open(position_files + pos_info, 'r')
            info_content = info_file.read().split('\n')

            amount_truncated = 0
            
            for index in range(len(info_content)):
                if len(info_content[index].split('-')) < 2:
                    continue
                coordinates = info_content[index].split('-')[1].split('*')
                y1, x1 = int(coordinates[0]), int(coordinates[1])
                
                
                if x1 > xmin or y1 > ymin:
                    amount_truncated += 1
            
            curr_dist = joint_frame_dist[j]
            
            for i in range(len(cf.BEHAVIOR_NAMES)):
                if i == 4:
                    curr_dist[i] = amount_truncated / cf.IMAGES_PER_INTERVAL * 1.0
                else:
                    curr_dist[i] *= (1 - amount_truncated / cf.IMAGES_PER_INTERVAL) * 1.0

            curr_dist = curr_dist / np.linalg.norm(curr_dist, ord = 1)
            ret.append( curr_dist )
        #print(len(ret))
        return ret
            
    
    def _time_shift(dist, length, int_len = cf.INTERVAL_LENGTH):
        """ Input and output array: [ [phase_len, phase_behavior, phase_start_interval, phase_end_interval] ] 
            Shifts any behavioral sequence by the rolling average order
        """
        if len(dist) <= 1 or length <= 0:
            return dist
        
        dist[0][0] -= length*int_len
        dist[0][3] -= length
        
        for j in range(1, len(dist)):
            dist[j][2] -= length
            dist[j][3] -= length
        
        dist[-1][0] += length*int_len
        dist[-1][3] += length
        
        return dist
        
        
    # read single frame file and merge it to time intervals
    single_frame_dist = _read_csv(single_frame_csv)
    single_frame_dist = _mark_truncated_images_single(single_frame_dist)
    single_frame_dist = _apply_rolling_average( single_frame_dist, cf.ROLLING_AVERAGE_SINGLE_FRAMES, cf.ROLLING_AVERAGE_WEIGHTS )
    sf_cumulated = _single_frame_prediction_to_intervalprediction(single_frame_dist)
    
    # read time interval file and apply rolling average
    joint_image_dist = _read_csv(joint_interval_csv)
    joint_image_dist = _mark_truncated_images_joint(joint_image_dist)
    joint_image_dist = _apply_rolling_average(joint_image_dist, cf.ROLLING_AVERAGE_JOINT_IMAGES, cf.ROLLING_AVERAGE_WEIGHTS)
    # merge the two prediction types
    ensemble_dist = _calculate_joint_prediction( joint_image_dist, sf_cumulated[0:-1])
    ensemble_dist = _apply_rolling_average(ensemble_dist, cf.ROLLING_AVERAGE_ENSEMBLE, cf.ROLLING_AVERAGE_WEIGHTS)
    
    
    
    save_path_csv = output_folder_prediction + 'raw_csv/ensemble/'
    filename_csv = datum + '_' + individual_code + '_ensemble.csv'
    _write_prediction_csv(ensemble_dist, save_path_csv, filename_csv)
    
    # apply post processing rules
    ensemble_sparse = _sparse_encoding(ensemble_dist)
    ensemble_sparse = _extract_single_phases( ensemble_sparse )
    ensemble_sparse = _remove_out(ensemble_sparse)
    ensemble_sparse = _map_behaviors(ensemble_sparse)
    ensemble_sparse = _remove_truncated_images(ensemble_sparse)
    ensemble_sparse = _remove_short_phases(ensemble_sparse)
    
    ensemble_sparse = _time_shift(ensemble_sparse, length = cf.ROLLING_AVERAGE_ENSEMBLE - 1)
    
    phases_ordered = [ [x[0] for x in ensemble_sparse if x[1] == 0], 
                      [x[0] for x in ensemble_sparse if x[1] == 1],
                      [x[0] for x in ensemble_sparse if x[1] == 2],
                      [x[0] for x in ensemble_sparse if x[1] == 3],
                      [x[0] for x in ensemble_sparse if x[1] == 4]
                     ]
    sparse_postprocessed = _get_sparse_intervals_from_phases(ensemble_sparse)
    
    outputfolder = output_folder_prediction + 'final/'
    filename = datum + '_' + individual_code + '_statistics' 
    _write_xlsx_statistics(phases_ordered, ensemble_sparse, sparse_postprocessed, outputfolder, filename)
    
    save_path = outputfolder + datum + '_' + individual_code + '_timeline'
    title = datum + '_' + individual_code
    _draw_timeline(ensemble_sparse, save_path, title)
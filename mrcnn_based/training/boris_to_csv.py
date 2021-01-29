#!/usr/bin/env python3
# -*- coding: utf-8 -*-
__author__ = "Max Hahn-Klimroth"
__copyright__ = "Copyright 2020, M. Hahn-Klimroth, T. Kapetanopoulos, J. Gübert, P. Dierkes"
__credits__ = ["J. Gübert", "P. Dierkes"]
__license__ = "MIT"
__version__ = "1.0"
__status__ = "Development"


import os
import csv
from openpyxl import load_workbook
import numpy as np
from collections import Counter
import sys

"""
Translates a folder full of BORIS files into csv label files in one-hot encoding fashion per time-interval.

Output: csv-label file
Time_interval,Start_Frame,End_Frame,Standing,Lying,Sleeping,Out

Requirements:
- Each BORIS file has format "YYYY-MM-DD_Art_Zoo_*.xlsx"
- Input folder contains only boris xlsx-files (and maybe other data types)
"""

INPUT_FOLDER = "" # folder which contains the corresponding xlsx files produces by BORIS export function
OUTPUT_FOLDER = ""

INDIVIDUAL_NUMBERS = {}
# e.g.:
#INDIVIDUAL_NUMBERS = {'Zebra-male-1': 1, 'Zebra-juven': 2}


# maps the behaviour of the BORIS file (col F in current version) to a number required by the deep learning system
BEHAVIOR_MAPPING = {}

# Example
# BEHAVIOR_MAPPING = {
    # "Standing": 0,
    # "S-Eating": 0,
    # "Playing": 0,
    # "Social-Interaction":0,
    # "LHU": 1,
    # "L-Eating": 1,
    # "LHD": 2,    
    # "Out": 3,
    # "Foraging": 0,
    # "Lying-Head-Unknown": 1
    # }

BEHAVIOR_CODES = {0: "Standing", 1: "LHU", 2: "LHD", 3: "Out"}
INTERVAL_LENGTH = 7
EXTENSION = "_SUM-7s_pred"

# Configuration of BORIS files.
TIME_COL = "A"
SUBJECT_COL = "E"
BEHAVIOR_COL = "F"
ACTION_COL = "I"

def _get_date_from_filename(filename):
    """ Returns date (first part of filename) """
    return filename.split("/")[-1].split("_")[0]

def _get_species_and_zoo_from_filename(filename):
    """ Returns species and zoo (2nd + 3rd part of filename) """
    return filename.split("/")[-1].split("_")[1], filename.split("/")[-1].split("_")[2]

def _get_first_content_row(ws):
    """ Returns the number of the first relevant row of the worksheet """
    for j in range(1, ws.max_row+1):
        if ws[TIME_COL+str(j)].value == "Time" or ws[TIME_COL+str(j)].value == "time":
            return j+1

def _get_individual_lists(ws, ind_numbers = INDIVIDUAL_NUMBERS, behav_map = BEHAVIOR_MAPPING):
    """ returns {1 : [time, behav, start/stop] }"""
    start_row = _get_first_content_row(ws = ws)
    ret_dict = {}
    
    for j in range(start_row, ws.max_row + 1):
        if j % 100 == 0:
            print("******Processed "+str(j)+" of "+str(ws.max_row)+" actions.")
            
        if not ws[SUBJECT_COL+str(j)].value in ind_numbers:
            # z.B. Nacht wird rausgefiltert
            continue

        if not ws[BEHAVIOR_COL+str(j)].value in behav_map.keys():
            # z.B. Orte (Box 1) werden rausgefiltert
            print("WARNING: " + ws[BEHAVIOR_COL+str(j)].value + " unknown. ")
            continue
        
        curr_ind = ind_numbers[ws[SUBJECT_COL+str(j)].value]
        curr_time = int(float(ws[TIME_COL+str(j)].value))

   
        curr_behav = behav_map[ws[BEHAVIOR_COL+str(j)].value]
        
        if not curr_ind in ret_dict.keys():
            ret_dict[curr_ind] = [ [curr_time,curr_behav, ws[ACTION_COL+str(j)].value] ]
        else:
            ret_dict[curr_ind].append([curr_time,curr_behav, ws[ACTION_COL+str(j)].value])
        
    return ret_dict
                                   

def _clean_list(boris_seq):
    """ Deletes [x, 0, start], [x, 0, stop], [x, 0, start], [x, 0, stop]"""
    new_seq = []
    j = 0
    while j < (len(boris_seq) - 3):
        new_seq.append(boris_seq[j])
        if not (boris_seq[j][1] == boris_seq[j+1][1] and boris_seq[j][1] == boris_seq[j+2][1] and boris_seq[j][1] == boris_seq[j+3][1]):            
            j += 1
        else:
            j += 3
    #append the last three entries.
    while j <= len(boris_seq)-1:
        new_seq.append(boris_seq[j])
        j += 1

    return new_seq

def _check_sanity(boris_seq):
    j = 0
    while j < (len(boris_seq) - 1):
        if not (boris_seq[j][1] == boris_seq[j+1][1] and boris_seq[j][2] == "START" and boris_seq[j+1][2] == "STOP"):
            print("Error on frame: "+ str(j)  )
            return False
        j += 2
    return True


def _reduce_to_intervals(behavior_seq, int_len = INTERVAL_LENGTH):
    """ Input: List of behaviors.
        Output: List of behaviors on intervals - by majority
    """
    chunk_seq = [behavior_seq[i * int_len:(i + 1) * int_len] for i in range((len(behavior_seq) + int_len - 1) // int_len )]
    ret_list = [Counter(x).most_common()[0][0] for x in chunk_seq]

    return ret_list
    
def _create_behavior_sequences(filepath, ind_numbers = INDIVIDUAL_NUMBERS):
    """ Returns for each second and each individual, what the behavior is
    {animal_number: [sequence of behaviors]}
    """
    ret_dict = {}

    wb = load_workbook(filename = filepath, read_only=True)
    ws = wb.active

    start_row = _get_first_content_row(ws = ws)
    len_vid = int(float(ws[TIME_COL+str(ws.max_row)].value))

    individual_lists = _get_individual_lists(ws)

    wb.close()
    
    for individual_code in individual_lists.keys():
        print("**** Cleaning list for individual" + str(individual_code))
        boris_sequence = _clean_list(individual_lists[individual_code])
        if not _check_sanity(boris_sequence):
            continue
        
        ret_seq = []

        curr_behav = boris_sequence[0][1]
        curr_time = 0
        
        for j in range(1, len(boris_sequence)):
            new_time = boris_sequence[j][0]
            time_diff = new_time - curr_time

            append_list = [curr_behav]*time_diff
            ret_seq.extend(append_list)
            
            curr_behav = boris_sequence[j][1]
            curr_time = boris_sequence[j][0]
            
        ret_dict[individual_code] = ret_seq
        
            
    return ret_dict        
        
    
def _extend_to_csv_rows(interval_sequence, int_len = INTERVAL_LENGTH, behav_codes = BEHAVIOR_CODES):
    """
    Input: List of behaviors on intervals.
    Output: List of Lists like: Time_interval,Start_Frame,End_Frame,Aktiv,Liegen,Schlafen,Out,Laufen
    """
    ret_list = []
    j = 1
    for behav in interval_sequence:
        tmp = [j, (j-1)*int_len+1, j*int_len]
        one_hot = list(map(int, np.zeros(len(behav_codes))))
        one_hot[behav] = 1
        tmp.extend(one_hot)
        ret_list.append(tmp)
        j += 1
    return ret_list
        
def _get_csv_titlerow(behav_codes = BEHAVIOR_CODES ):
    tmp = ["Time_interval", "Start_Frame", "End_Frame"]
    for behav in sorted(behav_codes.keys()):
        tmp.append(behav_codes[behav])
    return tmp

def _convert_one_boris_file(filepath = "", ind_numbers = INDIVIDUAL_NUMBERS, output_folder = OUTPUT_FOLDER, ext = EXTENSION):
    print("Starting to process "+filepath)
    if not os.path.isfile(filepath):
        print("No such file: "+filepath)
        return

    if not filepath.endswith(".xlsx"):
        return

    species, zoo =  _get_species_and_zoo_from_filename(filepath)
    base_filename = _get_date_from_filename(filepath)+ "_" + species + "_" + zoo

    print("**Create behavior sequence.")
    behavior_sequences = _create_behavior_sequences(filepath, ind_numbers = ind_numbers)

    for ind_num in behavior_sequences.keys():
        print("**Process interval mappings and prepare CSV for individual " + str(ind_num))
        interval_sequence = _reduce_to_intervals(behavior_sequences[ind_num])
        extended_rows = _extend_to_csv_rows(interval_sequence)
        
        output_path = output_folder + base_filename + "_" + str(ind_num) + ext + ".csv"
        
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)

        

        if sys.version_info >= (3, 0):
            with open(output_path, 'w+', newline='') as csv_out :
                writer = csv.writer(csv_out)
                writer.writerow(_get_csv_titlerow())
                writer.writerows(extended_rows)
        else:
            with open(output_path, 'wb+') as csv_out :
                writer = csv.writer(csv_out)
                writer.writerow(_get_csv_titlerow())
                writer.writerows(extended_rows)
        print("Created "+ output_path)

def convert_whole_folder(folder_path = INPUT_FOLDER, ind_numbers = INDIVIDUAL_NUMBERS, output_folder = OUTPUT_FOLDER, ext = EXTENSION):
    if not os.path.isdir(folder_path):
        return

    for boris_xlsx in os.listdir(folder_path):
        _convert_one_boris_file(folder_path+boris_xlsx)

if __name__ == "__main__":   
	convert_whole_folder()
